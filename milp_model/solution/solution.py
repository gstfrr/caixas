# -*- coding: utf-8 -*-
"""Solution File

This file is used to create and define useful functions to help the analysis of the solution. Here we can find
functions to calculate statistics and write files.
"""
from timeit import default_timer as timer

import pandas as pd
from openpyxl import load_workbook

from domain.data_dictionary import DataDictionary


def write_caixa_onda(caixa_onda_vars: DataDictionary) -> pd.DataFrame:
    all_vars = list(caixa_onda_vars.final_values())
    all_vars.sort(key=lambda x: x.onda.numero)
    rows = []
    for t in all_vars:
        if t.variable.index <= 0 or t.variable.X == 0:
            continue
        row = [t.onda.numero, t.caixa.caixa_id]
        rows.append(row)

    df = pd.DataFrame(rows, columns=['Onda', 'Caixa'])

    return df


def write_item_onda(item_onda_vars: DataDictionary) -> pd.DataFrame:
    all_vars = list(item_onda_vars.final_values())
    all_vars.sort(key=lambda x: x.onda.numero)
    rows = []
    for t in all_vars:
        if t.variable.index <= 0 or t.variable.X == 0:
            continue
        row = [t.onda.numero, t.item.nome]
        rows.append(row)

    df = pd.DataFrame(rows, columns=['Onda', 'Item'])

    return df


def write_output(solver, filename):
    print(f'\n\n\tOutput writing started to {filename}...')
    start = timer()

    df_caixa_onda = write_caixa_onda(solver.caixa_onda_vars)
    df_item_onda = write_item_onda(solver.item_onda_vars)

    with pd.ExcelWriter(filename) as writer:
        df_caixa_onda.to_excel(writer, sheet_name="Caixa-Onda", index=False)
        df_item_onda.to_excel(writer, sheet_name="Item-Onda", index=False)

    wb = load_workbook(filename)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    end = timer()
    print(f'\tOutput written to {filename} in {end - start} seconds')
